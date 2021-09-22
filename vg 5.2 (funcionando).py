import pandas as pd
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebElement
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import Cell
import time

df = pd.read_excel(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
planilha = openpyxl.load_workbook(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
folha = planilha['GERAL']

for i, consulta in enumerate(df['CPF']):
    cpfs = df.loc[i, "CPF"]
    statuses = df.loc[i, "STATUS"]
    link = "http://192.168.30.51/adapter/#/comercial/clientes"
    navegador = webdriver.Chrome(r'C:\Users\MOB\Documents\AAAAA VG\chromedriver.exe')
    navegador.get(link)
    navegador.implicitly_wait(10)
    navegador.find_element_by_xpath('//*[@id="login"]').send_keys("victor.castro")
    navegador.find_element_by_xpath('//*[@id="senha"]').send_keys("P@$$w0rd")
    navegador.find_element_by_xpath('//*[@id="senha"]').send_keys(Keys.ENTER)
    time.sleep(10)
    navegador.find_element_by_xpath('//*[@id="side-nav"]/li[2]/a').click()
    navegador.find_element_by_xpath('//*[@id="2-collapse"]/li[1]/a/span').click()
    time.sleep(10)
    navegador.find_element_by_xpath('//*[@id="content-container"]/div/div/section/div/form/fieldset/div[2]/div[2]/input').send_keys(f"{cpfs}")

    consulta0 = navegador.find_elements_by_xpath('//*[@id="content-container"]/div/div/section/div/form/fieldset/div[10]/button')
    consulta0

    if len(consulta0) > 0:
        navegador.find_element_by_xpath('//*[@id="content-container"]/div/div/section/div/form/fieldset/div[10]/button').click()
        time.sleep(10)
        consulta1 = navegador.find_elements_by_xpath('//*[@id="viewClientes"]/tbody/tr/td[4]/div/a')
        consulta1

        if len(consulta1) > 0 and consulta1[0].is_displayed():
            navegador.find_element_by_xpath('//*[@id="viewClientes"]/tbody/tr/td[4]/div/a').click()
            time.sleep(15)
            webdriver.ActionChains(navegador).send_keys(Keys.ESCAPE).perform()
            time.sleep(5)
            webdriver.ActionChains(navegador).send_keys(Keys.ESCAPE).perform()
            time.sleep(5)
            navegador.find_element_by_xpath('//*[@id="content-container"]/div[1]/ul/li[2]').click()
            time.sleep(10)
            consulta2 = navegador.find_elements_by_partial_link_text('HABILITADO')
            consulta2
            consulta3 = navegador.find_elements_by_partial_link_text('AGUARDANDO')
            consulta3
            consulta4 = navegador.find_elements_by_partial_link_text('CANCELADO')
            consulta4

     
            if len(consulta2) > 0 and consulta2[0].is_displayed():
                folha.cell(row=folha.max_row+1, column=4).value='HABILITADA'
                planilha.save(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
                print('HABILITADO')
                navegador.close()
        
            else: 
            
                if len(consulta3) > 0 and consulta3[0].is_displayed():
                    folha.cell(row=folha.max_row+1, column=4).value='AGUARDANDO'
                    planilha.save(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
                    print('AGUARDANDO')  
                    navegador.close()

                else:
                
                    if len(consulta4) > 0 and consulta4[0].is_displayed():
                        folha.cell(row=folha.max_row+1, column=4).value='CANCELADA'
                        planilha.save(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
                        print('CANCELADO')
                        navegador.close()

                    else:
                        folha.cell(row=folha.max_row+1, column=4).value='STATUS N/A'
                        planilha.save(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
                        print('STATUS NÃO ENCONTRADO')
                        navegador.close()           
    
        else:
            folha.cell(row=folha.max_row+1, column=4).value='CANCELADA'
            planilha.save(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
            print('CANCELADA')
            navegador.close()
    
    else:
        folha.cell(row=folha.max_row+1, column=4).value='CPF N/A'
        planilha.save(r'C:\Users\MOB\Documents\AAAAA VG\dados.xlsx')
        print('CLIENTE SEM CPF CADASTRADO')
        navegador.close()
        
# obg vgzinho szszsz